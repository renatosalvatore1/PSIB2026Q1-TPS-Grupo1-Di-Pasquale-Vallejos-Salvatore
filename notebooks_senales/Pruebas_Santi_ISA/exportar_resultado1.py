import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
def exportar_ISA(resultado_ISA):
    df_isa = pd.DataFrame(resultado_ISA)
    # columnas: sujeto, grupo, region, canal, mediana_isa

    wb = Workbook()
    wb.remove(wb.active)

    for region in df_isa["region"].unique():
        df_reg = df_isa[df_isa["region"] == region]

        # promedio por sujeto sobre los canales de la región
        df_sujeto = df_reg.groupby(["sujeto", "grupo"])["mediana_isa"].mean().reset_index()

        exp = df_sujeto[df_sujeto["grupo"] == "EXP"]["mediana_isa"].values
        ctr = df_sujeto[df_sujeto["grupo"] == "CTR"]["mediana_isa"].values

        ws = wb.create_sheet(title=region)
        ws["A1"], ws["B1"] = "EXP", "CTR"
        for i, val in enumerate(exp, start=2):
            ws.cell(row=i, column=1, value=val)
        for i, val in enumerate(ctr, start=2):
            ws.cell(row=i, column=2, value=val)

    wb.save("ISA_graphpad.xlsx")



def exportar_PSD(resultado_PSD):

    df = pd.DataFrame(resultado_PSD)

    wb = Workbook()
    wb.remove(wb.active)

    regiones = ['frontal', 'central', 'parietal', 'occipital']
    bandas = ['delta', 'theta', 'alfa', 'beta']

    for region in regiones:
        for banda in bandas:
            df_sub = df[(df['region'] == region) & (df['banda'] == banda)]

            # promedio por sujeto sobre los canales de la región/banda
            df_sujeto = df_sub.groupby(["sujeto", "grupo"])["potencia"].mean().reset_index()

            exp = df_sujeto[df_sujeto['grupo'] == 'EXP']['potencia'].values
            ctr = df_sujeto[df_sujeto['grupo'] == 'CTR']['potencia'].values

            ws = wb.create_sheet(title=f"{region}_{banda}")

            for col, label in enumerate(['EXP', 'CTR'], start=1):
                cell = ws.cell(row=1, column=col, value=label)
                cell.font = Font(bold=True, name='Arial')
                cell.fill = PatternFill('solid', start_color='D9E1F2')
                cell.alignment = Alignment(horizontal='center')

            for i, val in enumerate(exp, start=2):
                ws.cell(row=i, column=1, value=val).font = Font(name='Arial')
            for i, val in enumerate(ctr, start=2):
                ws.cell(row=i, column=2, value=val).font = Font(name='Arial')

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18

    wb.save('PSD_graphpad.xlsx')

def exportar_aperiodico(resultado_aperiodico):
    df = pd.DataFrame(resultado_aperiodico)

    wb = Workbook()
    wb.remove(wb.active)

    for region in ['frontal', 'central', 'parietal', 'occipital']:
        df_reg = df[df['region'] == region]

        # promedio por sujeto sobre los canales de la región
        df_sujeto = df_reg.groupby(["sujeto", "grupo"])["aperiodico"].mean().reset_index()


        exp = df_sujeto[df_sujeto['grupo'] == 'EXP']['aperiodico'].values
        ctr = df_sujeto[df_sujeto['grupo'] == 'CTR']['aperiodico'].values

        ws = wb.create_sheet(title=region)

        for col, label in enumerate(['EXP', 'CTR'], start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = Font(bold=True, name='Arial')
            cell.fill = PatternFill('solid', start_color='D9E1F2')
            cell.alignment = Alignment(horizontal='center')

        for i, val in enumerate(exp, start=2):
            ws.cell(row=i, column=1, value=val).font = Font(name='Arial')
        for i, val in enumerate(ctr, start=2):
            ws.cell(row=i, column=2, value=val).font = Font(name='Arial')

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18

    wb.save('aperiodico_graphpad.xlsx')


def exportar_asimetria(resultado_asimetria):
    df = pd.DataFrame(resultado_asimetria)

    wb = Workbook()
    wb.remove(wb.active)

    # promedio por sujeto sobre los canales de la región
    df_sujeto = df.groupby(["sujeto", "grupo"])["asimetria"].mean().reset_index()

    exp = df_sujeto[df_sujeto['grupo'] == 'EXP']['asimetria'].values
    ctr = df_sujeto[df_sujeto['grupo'] == 'CTR']['asimetria'].values

    ws = wb.create_sheet(title='FAA')

    for col, label in enumerate(['EXP', 'CTR'], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True, name='Arial')
        cell.fill = PatternFill('solid', start_color='D9E1F2')
        cell.alignment = Alignment(horizontal='center')

    for i, val in enumerate(exp, start=2):
        ws.cell(row=i, column=1, value=val).font = Font(name='Arial')
    for i, val in enumerate(ctr, start=2):
        ws.cell(row=i, column=2, value=val).font = Font(name='Arial')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18

    wb.save('asimetria_graphpad.xlsx')