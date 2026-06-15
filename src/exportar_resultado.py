import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'resultados_para_analisis')

def _get_path(filename):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return os.path.join(OUTPUT_DIR, filename)


def exportar_ISA(resultados_isa):
    df_isa = pd.DataFrame(resultados_isa)

    wb = Workbook()
    wb.remove(wb.active)

    for region in df_isa["region"].unique():
        df_reg = df_isa[df_isa["region"] == region]
        exp = df_reg[df_reg["grupo"] == "EXP"]["mediana"].values
        ctr = df_reg[df_reg["grupo"] == "CTR"]["mediana"].values

        ws = wb.create_sheet(title=region)

        for col, label in enumerate(['EXP', 'CTR'], start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = Font(bold=True, name='Arial')
            cell.fill = PatternFill('solid', start_color='D9E1F2')
            cell.alignment = Alignment(horizontal='center')

        for i, val in enumerate(exp, start=2):
            ws.cell(row=i, column=1, value=val).font = Font(name='Arial')
        for i, val in enumerate(ctr, start=2):
            ws.cell(row=i, column=2, value=val).font = Font(name='Arial')

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18

    wb.save(_get_path('ISA_graphpad.xlsx'))


def exportar_PSD(resultado_PSD):
    df = pd.DataFrame(resultado_PSD)
    df = df.groupby(["sujeto", "grupo", "region", "banda"])["potencia"].mean().reset_index()
    df["potencia"] = np.log10(df["potencia"])

    wb = Workbook()
    wb.remove(wb.active)

    for region in ['frontal', 'central', 'parietal', 'occipital']:
        for banda in ['delta', 'theta', 'alfa', 'beta']:
            df_sub = df[(df['region'] == region) & (df['banda'] == banda)]
            exp = df_sub[df_sub['grupo'] == 'EXP']['potencia'].values
            ctr = df_sub[df_sub['grupo'] == 'CTR']['potencia'].values

            ws = wb.create_sheet(title=f"{region}_{banda}")

            for col, label in enumerate(['EXP', 'CTR'], start=1):
                cell = ws.cell(row=1, column=col, value=label)
                cell.font = Font(bold=True, name='Arial')
                cell.fill = PatternFill('solid', start_color='D9E1F2')
                cell.alignment = Alignment(horizontal='center')

            for i, val in enumerate(exp, start=2):
                ws.cell(row=i, column=1, value=val).font = Font(name='Arial')
            for i, val in enumerate(ctr, start=2):
                ws.cell(row=i, column=2, value=val).font = Font(name='Arial')

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18

    wb.save(_get_path('PSD_graphpad.xlsx'))


def exportar_aperiodico(resultado_aperiodico):
    df = pd.DataFrame(resultado_aperiodico)
    df = df.groupby(["sujeto", "grupo", "region"])["aperiodico"].mean().reset_index()

    wb = Workbook()
    wb.remove(wb.active)

    for region in ['frontal', 'central', 'parietal', 'occipital']:
        df_reg = df[df['region'] == region]
        exp = df_reg[df_reg['grupo'] == 'EXP']['aperiodico'].values
        ctr = df_reg[df_reg['grupo'] == 'CTR']['aperiodico'].values

        ws = wb.create_sheet(title=region)

        for col, label in enumerate(['EXP', 'CTR'], start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = Font(bold=True, name='Arial')
            cell.fill = PatternFill('solid', start_color='D9E1F2')
            cell.alignment = Alignment(horizontal='center')

        for i, val in enumerate(exp, start=2):
            ws.cell(row=i, column=1, value=val).font = Font(name='Arial')
        for i, val in enumerate(ctr, start=2):
            ws.cell(row=i, column=2, value=val).font = Font(name='Arial')

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18

    wb.save(_get_path('aperiodico_graphpad.xlsx'))


def exportar_asimetria(resultado_asimetria):
    df = pd.DataFrame(resultado_asimetria)

    wb = Workbook()
    wb.remove(wb.active)

    exp = df[df['grupo'] == 'EXP']['asimetria'].values
    ctr = df[df['grupo'] == 'CTR']['asimetria'].values

    ws = wb.create_sheet(title='FAA')

    for col, label in enumerate(['EXP', 'CTR'], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True, name='Arial')
        cell.fill = PatternFill('solid', start_color='D9E1F2')
        cell.alignment = Alignment(horizontal='center')

    for i, val in enumerate(exp, start=2):
        ws.cell(row=i, column=1, value=val).font = Font(name='Arial')
    for i, val in enumerate(ctr, start=2):
        ws.cell(row=i, column=2, value=val).font = Font(name='Arial')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18

    wb.save(_get_path('asimetria_graphpad.xlsx'))